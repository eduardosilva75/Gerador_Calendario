import tkinter as tk
from tkinter import ttk, messagebox, colorchooser
import calendar
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

class CalendarGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("Gerador de Calendário com Ciclos de Folgas")
        self.root.geometry("900x700")
        self.root.resizable(True, True)  # Permite redimensionar
        
        # Configurações padrão
        self.ano = 2026
        self.ciclo_inicial = 1
        self.feriado_municipal = None  # (dia, mês)
        self.carnaval = None  # (dia, mês)
        self.ciclos = [
            [5, 6],  # S/D
            [3, 4],  # 5/6
            [2, 3],  # 4/5
            [1, 2],  # 3/4
            [0, 1],  # 2/3
            [0, 6],  # 2/D
        ]
        self.cores = {
            'fim_semana': 'FFF2CC',
            'feriado': 'FFCCC9',
            'folga': 'C9DAF8',
            'cabecalho': 'D9EAD3',
            'normal': 'FFFFFF'
        }
        
        self.criar_interface()
    
    def criar_interface(self):
        # Frame principal com duas colunas
        self.root.columnconfigure(0, weight=1)
        self.root.columnconfigure(1, weight=0)
        self.root.rowconfigure(0, weight=1)
        
        # Coluna esquerda - configurações (com scroll)
        canvas = tk.Canvas(self.root, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=canvas.yview)
        
        left_frame = ttk.Frame(canvas, padding="10")
        
        canvas.create_window((0, 0), window=left_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(10, 0), pady=10)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S), pady=10)
        
        # Atualizar scroll region quando o conteúdo mudar
        def on_frame_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        left_frame.bind("<Configure>", on_frame_configure)
        
        # Título
        titulo = ttk.Label(left_frame, text="Gerador de Calendário", 
                          font=('Calibri', 16, 'bold'))
        titulo.grid(row=0, column=0, columnspan=3, pady=10)
        
        # Ano
        ttk.Label(left_frame, text="Ano:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.ano_var = tk.IntVar(value=2026)
        ano_spin = ttk.Spinbox(left_frame, from_=2024, to=2030, textvariable=self.ano_var, width=10,
                               command=self.atualizar_nome_ficheiro)
        ano_spin.grid(row=1, column=1, sticky=tk.W, pady=5)
        self.ano_var.trace('w', lambda *args: self.atualizar_nome_ficheiro())
        
        # Ciclo inicial
        ttk.Label(left_frame, text="Ciclo inicial (1-6):").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.ciclo_var = tk.IntVar(value=1)
        self.ciclo_spin = ttk.Spinbox(left_frame, from_=1, to=6, textvariable=self.ciclo_var, width=10,
                                      command=self.atualizar_nome_ficheiro)
        self.ciclo_spin.grid(row=2, column=1, sticky=tk.W, pady=5)
        self.ciclo_var.trace('w', lambda *args: self.atualizar_nome_ficheiro())
        
        # Carnaval (Terça-feira)
        ttk.Label(left_frame, text="Carnaval:").grid(row=3, column=0, sticky=tk.W, pady=5)
        carnaval_frame = ttk.Frame(left_frame)
        carnaval_frame.grid(row=3, column=1, columnspan=2, sticky=tk.W, pady=5)
        
        self.carnaval_dia_var = tk.IntVar(value=17)
        carnaval_dia = ttk.Spinbox(carnaval_frame, from_=1, to=31, textvariable=self.carnaval_dia_var, width=5)
        carnaval_dia.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(carnaval_frame, text="/").pack(side=tk.LEFT, padx=2)
        
        self.carnaval_mes_var = tk.IntVar(value=2)
        carnaval_mes = ttk.Spinbox(carnaval_frame, from_=1, to=12, textvariable=self.carnaval_mes_var, width=5)
        carnaval_mes.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(carnaval_frame, text="(Terça-feira)", foreground='gray').pack(side=tk.LEFT, padx=5)
        
        # Feriado Municipal
        ttk.Label(left_frame, text="Feriado Municipal:").grid(row=4, column=0, sticky=tk.W, pady=5)
        feriado_frame = ttk.Frame(left_frame)
        feriado_frame.grid(row=4, column=1, columnspan=2, sticky=tk.W, pady=5)
        
        self.feriado_dia_var = tk.IntVar(value=14)
        feriado_dia = ttk.Spinbox(feriado_frame, from_=1, to=31, textvariable=self.feriado_dia_var, width=5)
        feriado_dia.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(feriado_frame, text="/").pack(side=tk.LEFT, padx=2)
        
        self.feriado_mes_var = tk.IntVar(value=6)
        feriado_mes = ttk.Spinbox(feriado_frame, from_=1, to=12, textvariable=self.feriado_mes_var, width=5)
        feriado_mes.pack(side=tk.LEFT, padx=2)
        
        ttk.Label(feriado_frame, text="(ex: 14/6 Abrantes)", foreground='gray').pack(side=tk.LEFT, padx=5)
        
        # Separador
        ttk.Separator(left_frame, orient='horizontal').grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=15)
        
        # CORREÇÃO: Mudar o row de 4 para 6
        # Seção de ciclos
        ttk.Label(left_frame, text="Configuração dos Ciclos de Folgas", 
                font=('Calibri', 12, 'bold')).grid(row=6, column=0, columnspan=3, pady=10)  # MUDADO de row=4 para row=6
        
        # Frame para ciclos
        self.ciclos_frame = ttk.Frame(left_frame)
        self.ciclos_frame.grid(row=7, column=0, columnspan=3, pady=10)  # MUDADO de row=5 para row=7
        
        self.ciclo_widgets = []
        self.criar_campos_ciclos()
        
        # Botões de gestão de ciclos
        btn_frame = ttk.Frame(left_frame)
        btn_frame.grid(row=8, column=0, columnspan=3, pady=10)  # MUDADO de row=6 para row=8
        
        ttk.Button(btn_frame, text="+ Adicionar Ciclo", 
                command=self.adicionar_ciclo).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="- Remover Último", 
                command=self.remover_ciclo).pack(side=tk.LEFT, padx=5)       
        
        # CORREÇÃO: Mudar o row do separador final
        # Separador
        ttk.Separator(left_frame, orient='horizontal').grid(row=9, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=15)  # MUDADO de row=11 para row=9
        
        # CORREÇÃO: Mudar o row do nome do ficheiro
        # Nome do ficheiro
        ttk.Label(left_frame, text="Nome do ficheiro:").grid(row=10, column=0, sticky=tk.W, pady=5)  # MUDADO de row=12 para row=10
        self.nome_var = tk.StringVar(value=f"Calendario_{self.ano}_Ciclo_{self.ciclo_inicial}")
        nome_entry = ttk.Entry(left_frame, textvariable=self.nome_var, width=30)
        nome_entry.grid(row=10, column=1, columnspan=2, sticky=tk.W, pady=5)  # MUDADO de row=12 para row=10
        
        # Coluna direita - Botão grande para gerar
        right_frame = ttk.Frame(self.root, padding="20")
        right_frame.grid(row=0, column=2, sticky=(tk.N, tk.S, tk.E), padx=10, pady=10)
        
        # Botão gerar com ciclos
        generate_btn = tk.Button(right_frame, 
                                text="🗓️\nGerar\ncom Ciclos",
                                command=self.gerar_calendario,
                                font=('Calibri', 14, 'bold'),
                                bg='#4CAF50',
                                fg='white',
                                width=12,
                                height=6,
                                relief=tk.RAISED,
                                borderwidth=3,
                                cursor='hand2')
        generate_btn.pack(pady=10)
        
        # Botão gerar sem ciclos
        generate_no_cycles_btn = tk.Button(right_frame, 
                                          text="📅\nGerar\nsem Ciclos",
                                          command=self.gerar_calendario_sem_ciclos,
                                          font=('Calibri', 14, 'bold'),
                                          bg='#2196F3',
                                          fg='white',
                                          width=12,
                                          height=6,
                                          relief=tk.RAISED,
                                          borderwidth=3,
                                          cursor='hand2')
        generate_no_cycles_btn.pack(pady=10)
        
        # Preview info
        info_label = ttk.Label(right_frame, 
                              text="Configura os ciclos\nà esquerda e escolhe\no tipo de calendário",
                              font=('Calibri', 9),
                              foreground='gray',
                              justify=tk.CENTER)
        info_label.pack(pady=10)
        
        # Botão fechar
        close_btn = tk.Button(right_frame, 
                             text="✕ Fechar",
                             command=self.root.quit,
                             font=('Calibri', 10),
                             bg='#f44336',
                             fg='white',
                             width=12,
                             height=2,
                             relief=tk.RAISED,
                             borderwidth=2,
                             cursor='hand2')
        close_btn.pack(pady=20, side=tk.BOTTOM)
    
    def criar_campos_ciclos(self):
        dias_semana = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
        
        # Limpar apenas widgets visuais (frames), não as variáveis
        for widget_set in self.ciclo_widgets:
            widget_set[0].destroy()  # Apenas o frame principal
        self.ciclo_widgets.clear()
        
        for i, ciclo in enumerate(self.ciclos):
            frame = ttk.Frame(self.ciclos_frame)
            frame.grid(row=i, column=0, pady=5, sticky=tk.W)
            
            label = ttk.Label(frame, text=f"Ciclo {i+1}:", width=10)
            label.pack(side=tk.LEFT, padx=5)
            
            dia1_var = tk.IntVar(value=ciclo[0])
            dia2_var = tk.IntVar(value=ciclo[1])
            
            combo1 = ttk.Combobox(frame, values=list(range(7)), width=8, state='readonly', 
                                 textvariable=dia1_var)
            combo1.current(ciclo[0])
            combo1.pack(side=tk.LEFT, padx=2)
            
            ttk.Label(frame, text="e").pack(side=tk.LEFT, padx=5)
            
            combo2 = ttk.Combobox(frame, values=list(range(7)), width=8, state='readonly',
                                 textvariable=dia2_var)
            combo2.current(ciclo[1])
            combo2.pack(side=tk.LEFT, padx=2)
            
            info = ttk.Label(frame, text=f"({dias_semana[ciclo[0]]} / {dias_semana[ciclo[1]]})", 
                           foreground='gray')
            info.pack(side=tk.LEFT, padx=10)
            
            self.ciclo_widgets.append([frame, dia1_var, dia2_var, combo1, combo2, info])
    
    def adicionar_ciclo(self):
        if len(self.ciclos) >= 12:
            messagebox.showwarning("Limite", "Máximo de 12 ciclos permitido!")
            return
        self.ciclos.append([0, 1])
        self.criar_campos_ciclos()
        self.ciclo_spin.config(to=len(self.ciclos))
    
    def remover_ciclo(self):
        if len(self.ciclos) <= 2:
            messagebox.showwarning("Limite", "Mínimo de 2 ciclos necessário!")
            return
        self.ciclos.pop()
        self.criar_campos_ciclos()
        self.ciclo_spin.config(to=len(self.ciclos))
    
    def atualizar_nome_ficheiro(self):
        """Atualiza automaticamente o nome do ficheiro quando ano ou ciclo mudam"""
        try:
            ano = self.ano_var.get()
            ciclo = self.ciclo_var.get()
            self.nome_var.set(f"Calendario_{ano}_Ciclo_{ciclo}")
        except:
            pass  # Ignora erros durante inicialização
    
    def criar_seletores_cor(self, parent):
        labels = {
            'fim_semana': 'Fins de semana:',
            'feriado': 'Feriados:',
            'folga': 'Folgas:',
            'cabecalho': 'Cabeçalhos:',
            'normal': 'Normal:'
        }
        
        for i, (key, label) in enumerate(labels.items()):
            ttk.Label(parent, text=label).grid(row=i, column=0, sticky=tk.W, padx=5, pady=2)
            
            cor_frame = tk.Frame(parent, bg=f'#{self.cores[key]}', width=40, height=20, 
                               relief=tk.SUNKEN, borderwidth=2)
            cor_frame.grid(row=i, column=1, padx=5, pady=2)
            cor_frame.grid_propagate(False)
    
    def atualizar_ciclos_da_interface(self):
        for i, widget_set in enumerate(self.ciclo_widgets):
            dia1_var = widget_set[1]
            dia2_var = widget_set[2]
            self.ciclos[i] = [dia1_var.get(), dia2_var.get()]
    
    def gerar_calendario(self):
        try:
            self.atualizar_ciclos_da_interface()
            
            ano = self.ano_var.get()
            ciclo_inicial = self.ciclo_var.get()
            nome_ficheiro = self.nome_var.get()
            
            if not nome_ficheiro:
                nome_ficheiro = f"Calendario_{ano}"
            
            if not nome_ficheiro.endswith('.xlsx'):
                nome_ficheiro += '.xlsx'
            
            # Gerar calendário com ciclos
            self.criar_calendario_excel(ano, ciclo_inicial, nome_ficheiro, com_ciclos=True)
            
            messagebox.showinfo("Sucesso", f"Calendário com ciclos gerado!\n{nome_ficheiro}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar calendário:\n{str(e)}")
    
    def gerar_calendario_sem_ciclos(self):
        try:
            ano = self.ano_var.get()
            nome_ficheiro = self.nome_var.get()
            
            # Ajustar nome se tiver "Ciclo"
            if "Ciclo" in nome_ficheiro:
                nome_ficheiro = nome_ficheiro.split("_Ciclo")[0]
            
            if not nome_ficheiro.endswith('.xlsx'):
                nome_ficheiro += '.xlsx'
            
            # Gerar calendário sem ciclos
            self.criar_calendario_excel(ano, 1, nome_ficheiro, com_ciclos=False)
            
            messagebox.showinfo("Sucesso", f"Calendário sem ciclos gerado!\n{nome_ficheiro}")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao gerar calendário:\n{str(e)}")
    
    def calcular_pascoa(self, ano):
        """Calcula a data da Páscoa usando o algoritmo de Meeus/Jones/Butcher"""
        a = ano % 19
        b = ano // 100
        c = ano % 100
        d = b // 4
        e = b % 4
        f = (b + 8) // 25
        g = (b - f + 1) // 3
        h = (19 * a + b - d - g + 15) % 30
        i = c // 4
        k = c % 4
        l = (32 + 2 * e + 2 * i - h - k) % 7
        m = (a + 11 * h + 22 * l) // 451
        mes = (h + l - 7 * m + 114) // 31
        dia = ((h + l - 7 * m + 114) % 31) + 1
        return date(ano, mes, dia)
    
    def criar_calendario_excel(self, ano, ciclo_inicial, nome_ficheiro, com_ciclos=True):
        # Calcular Páscoa e datas móveis
        pascoa = self.calcular_pascoa(ano)
        sexta_santa = pascoa - timedelta(days=2)
        corpo_deus = pascoa + timedelta(days=60)
        
        # Feriados fixos
        FERIADOS_PT = [
            date(ano, 1, 1),    # Ano Novo
            date(ano, 4, 25),   # Dia da Liberdade
            date(ano, 5, 1),    # Dia do Trabalhador
            date(ano, 6, 10),   # Dia de Portugal
            date(ano, 8, 15),   # Assunção
            date(ano, 10, 5),   # República
            date(ano, 11, 1),   # Todos os Santos
            date(ano, 12, 1),   # Restauração
            date(ano, 12, 8),   # Imaculada Conceição
            date(ano, 12, 25),  # Natal
        ]
        
        # Adicionar feriados móveis
        FERIADOS_PT.append(sexta_santa)  # Sexta-feira Santa
        FERIADOS_PT.append(pascoa)       # Domingo de Páscoa
        FERIADOS_PT.append(corpo_deus)   # Corpo de Deus
        
        # Adicionar Carnaval se definido
        try:
            dia_carnaval = self.carnaval_dia_var.get()
            mes_carnaval = self.carnaval_mes_var.get()
            if 1 <= dia_carnaval <= 31 and 1 <= mes_carnaval <= 12:
                carnaval = date(ano, mes_carnaval, dia_carnaval)
                if carnaval not in FERIADOS_PT:
                    FERIADOS_PT.append(carnaval)
        except:
            pass
        
        # Adicionar feriado municipal se definido
        try:
            dia_municipal = self.feriado_dia_var.get()
            mes_municipal = self.feriado_mes_var.get()
            if 1 <= dia_municipal <= 31 and 1 <= mes_municipal <= 12:
                feriado_municipal = date(ano, mes_municipal, dia_municipal)
                if feriado_municipal not in FERIADOS_PT:
                    FERIADOS_PT.append(feriado_municipal)
        except:
            pass  # Ignora se data inválida
        
        # Calcular folgas apenas se com_ciclos=True
        if com_ciclos:
            folgas = self.calcular_folgas_ano(ano, ciclo_inicial)
        else:
            folgas = set()  # Conjunto vazio - sem folgas
        
        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Calendário {ano}"
        
        # Estilos
        font_normal = Font(name='Calibri', size=12)
        font_head = Font(name='Calibri', size=13, bold=True)
        align_center = Alignment(horizontal='center', vertical='center')
        
        # Cabeçalho ano
        ws.merge_cells('A1:R1')
        ws['A1'] = str(ano)
        ws['A1'].font = Font(name='Calibri', size=18, bold=True)
        ws['A1'].alignment = align_center
        
        # Gerar meses
        meses_nome = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                      'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        meses = self.datas_por_mes(ano)
        
        row_start = 3
        col_left = 2
        col_right = 11
        
        for par in range(0, 12, 2):
            for i, mes_col in enumerate([col_left, col_right]):
                mes_idx = par + i
                if mes_idx >= 12:
                    break
                mes = meses_nome[mes_idx]
                semanas = meses[mes_idx + 1]
                
                # Nome do mês
                ws.merge_cells(start_row=row_start, start_column=mes_col, 
                             end_row=row_start, end_column=mes_col+6)
                ws.cell(row=row_start, column=mes_col).value = mes
                ws.cell(row=row_start, column=mes_col).font = font_head
                ws.cell(row=row_start, column=mes_col).fill = PatternFill('solid', 
                                                          fgColor=self.cores['cabecalho'])
                ws.cell(row=row_start, column=mes_col).alignment = align_center
                
                # Cabeçalho dias
                dias_sem = ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']
                for c, dia in enumerate(dias_sem):
                    ws.cell(row=row_start + 1, column=mes_col + c, value=dia)
                    ws.cell(row=row_start + 1, column=mes_col + c).font = font_normal
                    ws.cell(row=row_start + 1, column=mes_col + c).alignment = align_center
                
                # Dias do mês
                for l, semana in enumerate(semanas):
                    for c, data in enumerate(semana):
                        if data.month != mes_idx + 1:
                            continue
                        row = row_start + 2 + l
                        col = mes_col + c
                        ws.cell(row=row, column=col, value=data.day)
                        ws.cell(row=row, column=col).font = font_normal
                        ws.cell(row=row, column=col).alignment = align_center
                        
                        # Colorir
                        if data in FERIADOS_PT:
                            ws.cell(row=row, column=col).fill = PatternFill('solid', 
                                                              fgColor=self.cores['feriado'])
                        elif data in folgas:
                            ws.cell(row=row, column=col).fill = PatternFill('solid', 
                                                              fgColor=self.cores['folga'])
                        elif data.weekday() >= 5:
                            ws.cell(row=row, column=col).fill = PatternFill('solid', 
                                                              fgColor=self.cores['fim_semana'])
                        else:
                            ws.cell(row=row, column=col).fill = PatternFill('solid', 
                                                              fgColor=self.cores['normal'])
                
                # Número da semana
                for l, semana in enumerate(semanas):
                    primeira_data = next((d for d in semana if d.month == mes_idx + 1), None)
                    if primeira_data:
                        semana_iso = primeira_data.isocalendar()[1]
                        ws.cell(row=row_start + 2 + l, column=mes_col - 1, value=semana_iso)
                        ws.cell(row=row_start + 2 + l, column=mes_col - 1).font = font_normal
                        ws.cell(row=row_start + 2 + l, column=mes_col - 1).alignment = align_center
                
                # Largura colunas
                for c in range(mes_col - 1, mes_col + 7):
                    ws.column_dimensions[get_column_letter(c)].width = 8.5 if c >= mes_col else 3
                
                if i == 1:
                    row_start += len(semanas) + 3
        
        # Altura linhas
        for l in range(2, ws.max_row + 1):
            ws.row_dimensions[l].height = 26
        
        # Configurar para impressão em A4 - uma página centralizada
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToHeight = 1
        ws.page_setup.fitToWidth = 1
        ws.print_area = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
        
        # Centralizar na página
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True
        
        # Margens
        ws.page_margins.left = 0.6
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        
        wb.save(nome_ficheiro)
    
    def calcular_folgas_ano(self, ano, ciclo_inicio):
        folgas = set()
        dia_inicio = date(ano, 1, 5)
        semana_ref = dia_inicio.isocalendar()[1]
        
        data = date(ano, 1, 1)
        while data.year == ano:
            curr_semana = data.isocalendar()[1]
            ciclo_offset = (curr_semana - semana_ref) % len(self.ciclos)
            ciclo_pos = (ciclo_inicio - 1 + ciclo_offset) % len(self.ciclos)
            dias_folga = self.ciclos[ciclo_pos]
            if data.weekday() in dias_folga:
                folgas.add(data)
            data += timedelta(days=1)
        return folgas
    
    def datas_por_mes(self, ano):
        meses = {}
        for mes in range(1, 13):
            cal = calendar.Calendar(firstweekday=0)
            dias = list(cal.itermonthdates(ano, mes))
            semanas = [dias[i:i+7] for i in range(0, len(dias), 7)]
            meses[mes] = semanas
        return meses

if __name__ == "__main__":
    root = tk.Tk()
    app = CalendarGenerator(root)
    root.mainloop()